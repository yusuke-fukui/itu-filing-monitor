#!/usr/bin/env python3
"""
itu-filing-monitor / survey.py
ITU SpaceExplorer API から衛星ファイリング状態を一括取得し Excel に出力する。
スポット調査用ツール。config.yaml で調査条件を指定して実行する。

使い方:
  python3 src/survey.py --config survey_configs/ka_gso_japan_arc.yaml
  python3 src/survey.py --config survey_configs/leo_ngso_all.yaml
"""

import argparse
import json
import logging
import os
import sys
from datetime import date
from pathlib import Path

import requests
import yaml
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).parent.parent
LOG_DIR = ROOT / "logs"
LOG_DIR.mkdir(exist_ok=True)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s %(levelname)s %(message)s",
    handlers=[
        logging.FileHandler(LOG_DIR / "survey.log"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)

# ── API ──────────────────────────────────────────────────
BASE_URL = "https://www.itu.int/itu-r/space/apps/ep/spaceexplorer/v1/sns/dictionaries"
QUERIES_URL = "https://www.itu.int/itu-r/space/apps/ep/spaceexplorer/v1/queries"
PDF_BASE = "https://www.itu.int/sns/wic/"


# ── スタイル定数 ──────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", start_color=hex_color)

def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, name="Arial", size=size)

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

HDR_FILL   = _fill("1A5276")
HDR_FONT   = _font(bold=True, color="FFFFFF")
EVEN_FILL  = _fill("EBF5FB")
ODD_FILL   = _fill("FFFFFF")
SUM_FILL   = _fill("D6EAF8")
SUM_FONT   = _font(bold=True, color="1A5276")
BODY_FONT  = _font()
LINK_FONT  = Font(name="Arial", size=10, color="1155CC", underline="single")
BORDER     = _border()
CENTER     = Alignment(horizontal="center", vertical="center")
LEFT       = Alignment(horizontal="left",   vertical="center")


def _hdr(cell, text):
    cell.value = text
    cell.fill  = HDR_FILL
    cell.font  = HDR_FONT
    cell.alignment = CENTER
    cell.border = BORDER

def _body(cell, value, even=True, link=False, center=False):
    cell.value = value
    cell.fill  = EVEN_FILL if even else ODD_FILL
    cell.font  = LINK_FONT if link else BODY_FONT
    cell.alignment = CENTER if center else LEFT
    cell.border = BORDER


# ── 設定読み込み ───────────────────────────────────────────
def load_config(config_path: str) -> dict:
    with open(config_path) as f:
        cfg = yaml.safe_load(f)

    # config.yaml から Cookie を読み込む（存在する場合）
    config_yaml = Path(__file__).parent.parent / "config.yaml"
    if config_yaml.exists():
        with open(config_yaml) as f:
            main_cfg = yaml.safe_load(f) or {}
        if main_cfg.get("cookie") and not cfg.get("cookie"):
            cfg["cookie"] = main_cfg["cookie"]

    # デフォルト値
    cfg.setdefault("adm", [])  # 空リスト = 全国対象
    cfg.setdefault("orbit", {"geo": True, "ngso": True})
    cfg.setdefault("lon_from", None)
    cfg.setdefault("lon_to", None)
    cfg.setdefault("proc", {"api": True, "coord": True, "notification": True})
    cfg.setdefault("latest_brific", True)
    cfg.setdefault("take", 500)
    cfg.setdefault("label", "survey")
    cfg.setdefault("output", "survey_{label}_{date}.xlsx")

    # Cookie / 認証は config.yaml または環境変数から
    if not cfg.get("cookie"):
        cfg["cookie"] = os.environ.get("SPACEEXPLORER_COOKIE", "")

    return cfg


# ── APIセッション ─────────────────────────────────────────
def make_session(cookie: str) -> requests.Session:
    s = requests.Session()
    s.headers.update({
        "Accept": "*/*",
        "Content-Type": "application/json",
        "Referer": "https://www.itu.int/itu-r/space/apps/public/spaceexplorer/networks-explorer",
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36",
    })
    if cookie:
        s.headers["Cookie"] = cookie
    return s


# ── queries POST リクエスト構築 ────────────────────────────
def build_query_body(cfg: dict, skip: int = 0) -> dict:
    orbit = cfg["orbit"]
    proc  = cfg["proc"]

    # 周波数フィルタ（config に freq_min/freq_max があれば適用）
    freq_filter = []
    if cfg.get("freq_min") is not None or cfg.get("freq_max") is not None:
        freq_filter = [{
            "min": cfg.get("freq_min"),
            "max": cfg.get("freq_max"),
        }]

    body = {
        "frequencies": freq_filter,
        "general": {
            "network": "publication",
            "adm": cfg["adm"] if cfg["adm"] else [],
            "networkOrgCode": [],
            "satelliteName": cfg.get("satellite_name", []),
            "serviceType": "all",
        },
        "publications": {
            "latestBrIfic": cfg["latest_brific"],
            "brIficNumberFrom": None,
            "brIficNumberTo": None,
            "brIficDateFrom": None,
            "brIficDateTo": None,
        },
        "regDates": {
            "brNoticeDateReceiptFrom": None,
            "brNoticeDateReceiptTo": None,
            "networkProtectionDateFrom": None,
            "networkProtectionDateTo": None,
        },
        "regProcess": {
            "SuppressionStatus": "All",
            "masterReg": False,
            "nonPlanBand": proc.get("api", True) or proc.get("coord", True) or proc.get("notification", True),
            "bssPlanBand": False,
            "fssPlanBand": False,
            "esimPlanBand": False,
        },
        "satNetworkTypes": {
            "geo":      orbit.get("geo", True),
            "geoFrom":  cfg.get("lon_from"),
            "geoTo":    cfg.get("lon_to"),
            "nonGeo":   orbit.get("ngso", True),
            "apogeeFrom": None,
            "apogeeTo": None,
        },
        "skip": skip,
        "take": cfg["take"],
        "orderBy": "",
    }
    return body


# ── 全件取得（ページネーション対応）─────────────────────────
def fetch_all(session: requests.Session, cfg: dict) -> list[dict]:
    all_rows = []
    skip = 0
    take = cfg["take"]

    while True:
        body = build_query_body(cfg, skip=skip)
        log.info(f"  POST /queries skip={skip} take={take}")
        r = session.post(QUERIES_URL, json=body, timeout=60)
        r.raise_for_status()
        data = r.json()

        rows = data.get("table", [])
        total = data.get("total", 0)
        all_rows.extend(rows)

        log.info(f"  取得済み: {len(all_rows)} / {total}")

        if len(all_rows) >= total or len(rows) == 0:
            break
        skip += take

    return all_rows


# ── データ整形 ────────────────────────────────────────────
def parse_lon(long_nom) -> str:
    """long_nom（数値）を '110.0E' / '150.0W' 形式に変換"""
    if long_nom is None:
        return ""
    deg = float(long_nom)
    if deg >= 0:
        return f"{deg:.1f}E"
    else:
        return f"{abs(deg):.1f}W"

def proc_label(row: dict) -> str:
    """手続き種別を日本語ラベルで返す"""
    ssn = row.get("ssn_ref", "") or ""
    if row.get("type_api"):
        return f"API ({ssn})"
    if row.get("type_coord"):
        return f"CR ({ssn})"
    if row.get("type_notification"):
        return f"Notification ({ssn})"
    return ssn

def pdf_url(row: dict) -> str:
    path = row.get("pdf_path", "")
    if not path:
        return ""
    path = path.replace("\\", "/")
    return f"https://www.itu.int/sns/wic/{path}"


def parse_freq(row: dict, key: str) -> str:
    """freq_tooltip_emi / freq_tooltip_rcp を '694-960, 1710-2170 MHz' 形式に変換"""
    raw = row.get(key)
    if not raw:
        return ""
    try:
        data = json.loads(raw) if isinstance(raw, str) else raw
        ranges = data.get("ranges", [])
        if not ranges:
            return ""
        # 隣接・重複レンジを統合してサマリ表示
        parts = []
        for r in ranges[:8]:  # 最大8レンジまで（セル幅対応）
            f, t = r.get("f", 0), r.get("t", 0)
            if abs(t - f) < 0.01:
                parts.append(f"{f:.1f}")
            else:
                parts.append(f"{f:.1f}-{t:.1f}")
        result = ", ".join(parts)
        if len(ranges) > 8:
            result += f" ... (+{len(ranges)-8})"
        return result + " MHz"
    except Exception:
        return ""


# ── Excel 出力 ────────────────────────────────────────────
COLUMNS = [
    ("ADM",          8,  "adm"),
    ("Satellite Name", 24, "sat_name"),
    ("軌道位置",      12, "_lon"),
    ("手続き種別",    18, "_proc"),
    ("受理日",        12, "d_rcv"),
    ("BR IFIC No.",   10, "wic_no"),
    ("掲載日",        12, "d_wic"),
    ("保護期限",      12, "d_prot_eff_max"),
    ("BIU期限",       12, "d_inuse_max"),
    ("登録期限",      12, "d_reg_limit_max"),
    ("送信周波数 (MHz)", 36, "_freq_emi"),
    ("受信周波数 (MHz)", 36, "_freq_rcp"),
    ("PDF",           10, "_pdf"),
]

def write_sheet(ws, rows: list[dict], title: str):
    ws.title = title

    # ヘッダー
    for col, (hdr, width, _) in enumerate(COLUMNS, 1):
        _hdr(ws.cell(row=1, column=col), hdr)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    for r, row in enumerate(rows, 2):
        even = (r % 2 == 0)
        lon_str   = parse_lon(row.get("long_nom"))
        proc_str  = proc_label(row)
        pdf_str   = pdf_url(row)

        values = {
            "adm":            row.get("adm", ""),
            "sat_name":       (row.get("sat_name") or "").strip(),
            "_lon":           lon_str,
            "_proc":          proc_str,
            "d_rcv":          row.get("d_rcv", ""),
            "wic_no":         row.get("wic_no"),
            "d_wic":          row.get("d_wic", ""),
            "d_prot_eff_max": row.get("d_prot_eff_max", ""),
            "d_inuse_max":    row.get("d_inuse_max", ""),
            "d_reg_limit_max":row.get("d_reg_limit_max", ""),
            "_freq_emi":      parse_freq(row, "freq_tooltip_emi"),
            "_freq_rcp":      parse_freq(row, "freq_tooltip_rcp"),
            "_pdf":           pdf_str,
        }

        for col, (_, _, key) in enumerate(COLUMNS, 1):
            val = values[key]
            is_link = (key == "_pdf") and bool(val)
            is_center = key in ("adm", "_lon", "_proc", "d_rcv", "wic_no", "d_wic",
                                "d_prot_eff_max", "d_inuse_max", "d_reg_limit_max")
            if key in ("_freq_emi", "_freq_rcp"):
                cell.alignment = cell.alignment.copy(wrapText=True)
            cell = ws.cell(row=r, column=col)
            _body(cell, val, even=even, link=is_link, center=is_center)

            # PDF列はハイパーリンク
            if is_link:
                cell.hyperlink = val
                cell.value = "PDF"


def write_summary_sheet(ws, rows: list[dict], cfg: dict):
    ws.title = "サマリ"

    today = date.today().isoformat()
    label = cfg.get("label", "")

    ws["A1"] = f"ITU衛星ファイリング状態サマリ — {label}"
    ws["A1"].font = _font(bold=True, size=13, color="1A5276")
    ws["A2"] = f"取得日: {today}　　対象国: {', '.join(cfg['adm'])}　　データソース: ITU SpaceExplorer"
    ws["A2"].font = _font(size=10, color="7F8C8D")
    ws.row_dimensions[1].height = 24

    # 国 × 手続き種別 クロス集計
    adm_list = sorted(set(r["adm"] for r in rows))
    proc_cols = ["API", "CR", "Notification", "その他", "合計"]

    ws["A4"] = "ADM"
    _hdr(ws["A4"], "ADM")
    for i, p in enumerate(proc_cols, 2):
        _hdr(ws.cell(row=4, column=i), p)

    for ri, adm in enumerate(adm_list, 5):
        api_n  = sum(1 for r in rows if r["adm"] == adm and r.get("type_api"))
        cr_n   = sum(1 for r in rows if r["adm"] == adm and r.get("type_coord"))
        ntf_n  = sum(1 for r in rows if r["adm"] == adm and r.get("type_notification"))
        other  = sum(1 for r in rows if r["adm"] == adm and not any([r.get("type_api"), r.get("type_coord"), r.get("type_notification")]))
        total  = api_n + cr_n + ntf_n + other

        even = (ri % 2 == 0)
        for ci, val in enumerate([adm, api_n, cr_n, ntf_n, other, total], 1):
            _body(ws.cell(row=ri, column=ci), val, even=even, center=True)

    # 合計行
    total_row = len(adm_list) + 5
    ws.cell(row=total_row, column=1).value = "合計"
    for ci in range(1, 7):
        c = ws.cell(row=total_row, column=ci)
        if ci > 1:
            col_letter = get_column_letter(ci)
            c.value = f"=SUM({col_letter}5:{col_letter}{total_row - 1})"
        c.fill = SUM_FILL
        c.font = SUM_FONT
        c.alignment = CENTER
        c.border = BORDER

    for i in range(1, 7):
        ws.column_dimensions[get_column_letter(i)].width = 14

    # 凡例
    ws.cell(row=total_row + 2, column=1).value = "手続き種別"
    ws.cell(row=total_row + 2, column=1).font = _font(bold=True, size=11, color="1A5276")

    legend = [
        ("API",           "Advance Publication — 事前公表（Art.9.1）"),
        ("CR",            "Coordination Request — 調整要求（Art.9.6/9.7）"),
        ("Notification",  "Notification — 通知（Art.11）"),
        ("その他",        "Plan Band（AP30/AP30A/AP30B）等"),
    ]
    for li, (code, desc) in enumerate(legend, total_row + 3):
        ws.cell(row=li, column=1).value = code
        ws.cell(row=li, column=1).font = _font(bold=True)
        ws.cell(row=li, column=2).value = desc
        ws.cell(row=li, column=2).font = _font()
        ws.merge_cells(f"B{li}:F{li}")


def write_history_sheet(ws, rows: list[dict]):
    """衛星ごとの手続き変遷シート（latest_brific=False のとき有効）"""
    ws.title = "変遷"

    # ヘッダー
    hist_cols = [
        ("ADM",        8),
        ("Satellite Name", 24),
        ("手続き種別",  18),
        ("改版",       10),
        ("受理日",     12),
        ("BR IFIC No.", 10),
        ("掲載日",     12),
        ("手続きステップ", 20),
    ]
    for col, (h, w) in enumerate(hist_cols, 1):
        _hdr(ws.cell(row=1, column=col), h)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # sat_name ごとにグループ化して受理日順に並べる
    from collections import defaultdict
    groups: dict[str, list] = defaultdict(list)
    for r in rows:
        key = f"{r.get('adm','')}__{(r.get('sat_name') or '').strip()}"
        groups[key].append(r)

    # 手続きステップのラベル
    def step_label(row):
        ssn = row.get("ssn_ref", "") or ""
        if row.get("type_api"):      return f"① API ({ssn})"
        if row.get("type_coord"):    return f"② CR ({ssn})"
        if row.get("type_notification"): return f"③ Notification ({ssn})"
        return ssn

    r_idx = 2
    for key, group in sorted(groups.items()):
        group_sorted = sorted(group, key=lambda x: x.get("d_rcv") or "")
        for i, row in enumerate(group_sorted):
            even = (r_idx % 2 == 0)
            rev = row.get("pub_rev") or ""
            vals = [
                row.get("adm", ""),
                (row.get("sat_name") or "").strip(),
                row.get("ssn_ref", ""),
                rev,
                row.get("d_rcv", ""),
                row.get("wic_no"),
                row.get("d_wic", ""),
                step_label(row),
            ]
            for col, val in enumerate(vals, 1):
                _body(ws.cell(row=r_idx, column=col), val,
                      even=even, center=(col not in (2, 8)))
            r_idx += 1

        # 衛星ごとに区切り線（最終行の下に薄いボーダー）
        if group_sorted:
            sep = Side(style="medium", color="1A5276")
            for col in range(1, 9):
                c = ws.cell(row=r_idx - 1, column=col)
                c.border = Border(
                    left=c.border.left, right=c.border.right,
                    top=c.border.top,   bottom=sep
                )



def write_meta_sheet(ws, cfg: dict, config_path: str, rows: list):
    """調査条件・実行情報シート"""
    ws.title = "調査条件"
    from datetime import datetime

    # ── タイトル ──
    ws["A1"] = "ITU Filing Survey — 調査条件・実行記録"
    ws["A1"].font = _font(bold=True, size=13, color="1A5276")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:C1")

    def _row(ws, r, label, value, note=""):
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = _font(bold=True, color="1A5276", size=10)
        lc.fill = _fill("EBF5FB")
        lc.border = BORDER
        lc.alignment = LEFT

        vc = ws.cell(row=r, column=2, value=value)
        vc.font = _font(size=10)
        vc.border = BORDER
        vc.alignment = LEFT

        if note:
            nc = ws.cell(row=r, column=3, value=note)
            nc.font = _font(size=10, color="7F8C8D")
            nc.border = BORDER
            nc.alignment = LEFT

    r = 3
    ws.cell(row=r-1, column=1, value="■ 実行情報").font = _font(bold=True, size=11, color="1A5276")

    _row(ws, r,   "実行日時",       datetime.now().strftime("%Y-%m-%d %H:%M"),  "JST")
    _row(ws, r+1, "設定ファイル",   config_path,                                "")
    _row(ws, r+2, "調査ラベル",     cfg.get("label", ""),                       "")
    _row(ws, r+3, "取得総件数",     len(rows),                                  f"GSO: {sum(1 for x in rows if x.get('ntc_type')=='G')}  /  NGSO: {sum(1 for x in rows if x.get('ntc_type')=='N')}")
    _row(ws, r+4, "データソース",   "ITU SpaceExplorer",                        "https://www.itu.int/itu-r/space/apps/public/spaceexplorer/")

    r += 7
    ws.cell(row=r-1, column=1, value="■ 調査条件").font = _font(bold=True, size=11, color="1A5276")

    adm_str = ", ".join(cfg.get("adm", [])) if cfg.get("adm") else "指定なし（全国対象）"
    orbit = cfg.get("orbit", {})
    orbit_str = []
    if orbit.get("geo"):   orbit_str.append("GSO")
    if orbit.get("ngso"):  orbit_str.append("NGSO")

    lon_from = cfg.get("lon_from")
    lon_to   = cfg.get("lon_to")
    lon_str  = "指定なし（全軌道位置）"
    lon_note = ""
    if not orbit.get("geo") and orbit.get("ngso"):
        # NGSO のみの場合は軌道位置フィルタが無意味
        lon_str  = "— （NGSOのため適用なし）"
        lon_note = "NGSOはlong_nomがnullのため軌道位置フィルタは無効"
    elif lon_from is not None or lon_to is not None:
        lon_str  = f"{lon_from if lon_from is not None else '—'}° 〜 {lon_to if lon_to is not None else '—'}°"
        lon_note = "GSOの軌道位置（東経+/西経−）"
    else:
        lon_note = "指定なし → 全軌道位置を取得"

    freq_min = cfg.get("freq_min")
    freq_max = cfg.get("freq_max")
    freq_str = "指定なし（全周波数）"
    if freq_min is not None or freq_max is not None:
        freq_str = f"{freq_min if freq_min is not None else '—'} 〜 {freq_max if freq_max is not None else '—'} MHz"

    proc = cfg.get("proc", {})
    proc_list = [k.upper() for k, v in proc.items() if v]
    proc_str = ", ".join(proc_list) if proc_list else "すべて"

    brific_str = "最新のみ（Latest BR IFIC）" if cfg.get("latest_brific", True) else "全BR IFIC（変遷シートあり）"

    _row(ws, r,   "対象国 (ADM)",    adm_str,               "")
    _row(ws, r+1, "軌道種別",        " / ".join(orbit_str), "")
    _row(ws, r+2, "軌道位置フィルタ", lon_str,               lon_note)
    _row(ws, r+3, "周波数フィルタ",  freq_str,              "MHz 単位")
    _row(ws, r+4, "手続き種別",      proc_str,              "")
    _row(ws, r+5, "BR IFIC 対象",    brific_str,            "")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 30


def build_excel(rows: list[dict], cfg: dict, output_path: str):
    wb = Workbook()

    gso_rows  = [r for r in rows if r.get("ntc_type") == "G"]
    ngso_rows = [r for r in rows if r.get("ntc_type") == "N"]

    ws_gso = wb.active
    write_sheet(ws_gso, gso_rows, "GSO")

    ws_ngso = wb.create_sheet()
    write_sheet(ws_ngso, ngso_rows, "NGSO")

    ws_sum = wb.create_sheet()
    write_summary_sheet(ws_sum, rows, cfg)

    # latest_brific=False のときは変遷シートも出力
    if not cfg.get("latest_brific", True):
        ws_hist = wb.create_sheet()
        write_history_sheet(ws_hist, rows)
        log.info("変遷シートを追加しました（latest_brific=false）")

    ws_meta = wb.create_sheet(index=0)  # 先頭シートに配置
    write_meta_sheet(ws_meta, cfg, cfg.get("_config_path", ""), rows)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True); wb.save(output_path)
    log.info(f"Excel保存: {output_path}  (GSO:{len(gso_rows)} NGSO:{len(ngso_rows)} 合計:{len(rows)})")


# ── メイン ────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="ITU Filing Survey Tool")
    parser.add_argument("--filter", nargs="+", default=None, help="衛星名の部分一致フィルタ")
    parser.add_argument("--config", required=True, help="調査設定ファイル（yaml）")
    args = parser.parse_args()

    cfg = load_config(args.config)
    cfg["_config_path"] = args.config  # メタシート用

    label  = cfg["label"]
    today  = date.today().strftime("%Y%m%d")
    output = cfg["output"].replace("{label}", label).replace("{date}", today)

    log.info(f"=== ITU Filing Survey ===")
    log.info(f"設定: {args.config}")
    adm_label = ", ".join(cfg["adm"]) if cfg["adm"] else "全国（指定なし）"
    log.info(f"対象国: {adm_label}")
    log.info(f"軌道: geo={cfg['orbit'].get('geo')} ngso={cfg['orbit'].get('ngso')}")
    if cfg.get("lon_from") or cfg.get("lon_to"):
        log.info(f"軌道位置フィルタ: {cfg.get('lon_from')}° 〜 {cfg.get('lon_to')}°")

    if not cfg.get("cookie"):
        log.warning("SPACEEXPLORER_COOKIE が未設定です。認証なしで試みます。")

    session = make_session(cfg.get("cookie", ""))

    try:
        rows = fetch_all(session, cfg)
    except requests.HTTPError as e:
        log.error(f"API取得失敗: {e}")
        sys.exit(1)

    if args.filter:
        rows = [r for r in rows if any(k.upper() in (r.get("sat_name") or "").upper() for k in args.filter)]
        log.info(f"フィルタ '{args.filter}' 適用後: {len(rows)}件")

    if args.filter:
        rows = [r for r in rows if any(k.upper() in (r.get("sat_name") or "").upper() for k in args.filter)]
        log.info(f"フィルタ '{args.filter}' 適用後: {len(rows)}件")
    if not rows:
        log.warning("取得結果が0件でした。設定や認証を確認してください。")
        sys.exit(0)

    if args.filter:
        rows = [r for r in rows if any(k.upper() in (r.get("sat_name") or "").upper() for k in args.filter)]
        log.info(f"フィルタ '{args.filter}' 適用後: {len(rows)}件")
    if args.filter:
        rows = [r for r in rows if any(k.upper() in (r.get("sat_name") or "").upper() for k in args.filter)]
        log.info(f"フィルタ適用後: {len(rows)}件")
    build_excel(rows, cfg, output)
    log.info("Done.")


if __name__ == "__main__":
    main()
